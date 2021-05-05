from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#wb = Workbook()
#ws = wb.active

#w1 = wb.create_sheet('Newsheet')
#w2 = wb.create_sheet('Another', 0)

#ws.title = 'Mysheet'
#wb.save('new')

wb2 = load_workbook('regions.xlsx')

new_sheet = wb2.create_sheet('Newsheet')

active_sheet = wb2.active


cell =active_sheet['A1']

active_sheet['A1'] = 0
wb2.save('modified2.xlsx')
